from flask import Flask, request, jsonify, session, send_from_directory, redirect
from flask_cors import CORS
import psycopg2
import psycopg2.extras
import psycopg2.errors
import bcrypt
import os
import time
import uuid
import json
import random
import string
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, date, timedelta
from dotenv import load_dotenv
import secrets
import hashlib
import csv as _csv
import io as _io
from flask import Response as _Response
import requests
import jwt as pyjwt

load_dotenv()

app = Flask(__name__, static_folder='../frontend', static_url_path='')
app.secret_key = os.environ.get('SECRET_KEY', 'sistema-rh-secret-key-change-in-production')
CORS(app, supports_credentials=True)

DATABASE_URL = os.environ.get('DATABASE_URL')
UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), '..', 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ── SSO — Valore Hub ─────────────────────────────────────────────────────────
HUB_BASE_URL = os.environ.get('HUB_BASE_URL', 'https://portal.valore.com.br').rstrip('/')
SSO_CLIENT_ID = os.environ.get('SSO_CLIENT_ID', 'portal-rh')
SSO_CLIENT_SECRET = os.environ.get('SSO_CLIENT_SECRET', '')
RH_PROVISIONING_API_KEY = os.environ.get('RH_PROVISIONING_API_KEY', '')

# Mapeia o departamento cadastrado no RH para o nome do grupo no Hub (usado
# para liberar os softwares/ferramentas padrão daquele departamento). Grupos
# sem equivalente direto no Hub são criados lá com o mesmo nome do RH.
RH_DEPT_TO_HUB_GROUP = {
    'Contábil': 'Contábil',
    'Fiscal': 'Fiscal',
    'Departamento Pessoal': 'Departamento Pessoal',
    'Sucesso do Cliente': 'Sucesso do Cliente',
    'Administrativo': 'Administrativo',
    'Paralegal': 'Paralegal',
    'Diretoria': 'Diretoria',
    'Marketing': 'Marketing',
    'Tecnologia': 'Tecnologia',
    # Financeiro e Consultoria eram grupos duplicados no Hub — unificados em
    # "Consultoria", que é quem tem os acessos de verdade hoje.
    'BPO Financeiro': 'Consultoria',
    'Financeiro': 'Consultoria',
    'Consultoria': 'Consultoria',
}


def provisionar_usuario_hub(nome, email, departamento, cargo=''):
    """
    Cria automaticamente o usuário no Hub na admissão do funcionário, sem
    nenhum acesso a portal — só o grupo do departamento (que já libera os
    softwares/ferramentas padrão) e o cargo, que passa a ser preenchido só
    por aqui (não é mais editável manualmente na tela de gestão do Hub).
    Falha aberta: se o Hub estiver fora do ar, não bloqueia a admissão, só
    loga o problema.
    """
    if not email or '@' not in email or not RH_PROVISIONING_API_KEY:
        return
    try:
        requests.post(
            f'{HUB_BASE_URL}/api/provisioning/rh/',
            json={
                'api_key': RH_PROVISIONING_API_KEY,
                'name': nome or '',
                'email': email,
                'department_group': RH_DEPT_TO_HUB_GROUP.get(departamento or '', departamento or ''),
                'position': cargo or '',
            },
            timeout=10,
        )
    except requests.RequestException as exc:
        app.logger.warning('Falha ao provisionar usuário no Hub (email=%s): %s', email, exc)
SSO_DASHBOARD_PATH = os.environ.get('SSO_DASHBOARD_PATH', '/')

# Mapeia o perfil do Hub (RH_PORTAL_ROLE_CHOICES) para os valores de `role`
# já usados neste sistema (ver ROLE_LABELS no frontend/js/app.js).
MAPA_PERFIL_HUB_RH = {
    'rh_colaborador': 'colaborador',
    'rh_lideranca': 'lideranca',
    'rh_gestor': 'admin',
    'rh_administrador': 'admin',
}

_hub_public_key_cache = {'pem': None, 'fetched_at': 0.0}
_HUB_PUBLIC_KEY_CACHE_SECONDS = 3600


def obter_chave_publica_hub(forcar_atualizacao=False):
    agora = time.time()
    if not forcar_atualizacao and _hub_public_key_cache['pem'] and (agora - _hub_public_key_cache['fetched_at'] < _HUB_PUBLIC_KEY_CACHE_SECONDS):
        return _hub_public_key_cache['pem']
    resp = requests.get(f'{HUB_BASE_URL}/api/sso/public-key/', timeout=10)
    resp.raise_for_status()
    _hub_public_key_cache['pem'] = resp.text
    _hub_public_key_cache['fetched_at'] = agora
    return resp.text


class SSOClientError(Exception):
    pass


def trocar_codigo_por_claims(code):
    """Troca o código do Hub pelo id_token (RS256) e valida a assinatura, issuer e audience."""
    try:
        resp = requests.post(
            f'{HUB_BASE_URL}/api/sso/exchange/',
            json={'code': code, 'client_id': SSO_CLIENT_ID, 'client_secret': SSO_CLIENT_SECRET},
            timeout=10,
        )
    except requests.RequestException as exc:
        raise SSOClientError(f'Falha ao conectar ao Hub: {exc}')

    if resp.status_code != 200:
        raise SSOClientError(f'Hub recusou o código: {resp.text}')

    id_token = resp.json().get('id_token', '')

    try:
        claims = pyjwt.decode(id_token, obter_chave_publica_hub(), algorithms=['RS256'], audience=SSO_CLIENT_ID)
    except pyjwt.PyJWTError:
        try:
            claims = pyjwt.decode(id_token, obter_chave_publica_hub(forcar_atualizacao=True), algorithms=['RS256'], audience=SSO_CLIENT_ID)
        except pyjwt.PyJWTError as exc2:
            raise SSOClientError(f'Token do Hub inválido: {exc2}')

    if claims.get('iss') != HUB_BASE_URL:
        raise SSOClientError('Token emitido por um emissor não reconhecido.')
    if claims.get('portal') != SSO_CLIENT_ID:
        raise SSOClientError('Token emitido para outro portal.')

    return claims


def localizar_ou_criar_usuario_via_sso(claims):
    """Vincula (por hub_user_id) ou cria o usuário local a partir das claims do Hub."""
    hub_user_id = str(claims['sub'])
    email = claims['email'].strip().lower()
    nome = claims.get('name') or email.split('@')[0]
    portal_role = claims.get('portal_role') or 'consulta'
    role_local = MAPA_PERFIL_HUB_RH.get(portal_role, 'colaborador')

    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("SELECT * FROM users WHERE hub_user_id = %s", (hub_user_id,))
        row = c.fetchone()
        user = row_to_dict(row, c) if row else None

        if not user:
            c.execute("SELECT * FROM users WHERE email = %s", (email,))
            row = c.fetchone()
            user = row_to_dict(row, c) if row else None
            if user:
                c.execute("UPDATE users SET hub_user_id = %s WHERE id = %s", (hub_user_id, user['id']))
            else:
                novo_id = str(uuid.uuid4())
                senha_inutilizavel = bcrypt.hashpw(secrets.token_urlsafe(32).encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
                c.execute(
                    """INSERT INTO users (id, name, email, password_hash, role, hub_user_id, active)
                       VALUES (%s, %s, %s, %s, %s, %s, 1)""",
                    (novo_id, nome, email, senha_inutilizavel, role_local, hub_user_id),
                )
                user = {'id': novo_id, 'active': 1}

        if not user.get('active'):
            raise SSOClientError('Usuário está bloqueado neste portal.')

        c.execute(
            """UPDATE users SET name = %s, role = %s, hub_portal_role = %s, hub_session_version = %s,
               last_sso_login = %s, permission_checked_at = %s WHERE id = %s""",
            (nome, role_local, portal_role, claims.get('session_version'), datetime.utcnow(), datetime.utcnow(), user['id']),
        )
        conn.commit()

        c.execute("SELECT id, name, email, role, department FROM users WHERE id = %s", (user['id'],))
        row = c.fetchone()
        return row_to_dict(row, c)
    finally:
        conn.close()


def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    return conn


def row_to_dict(row, cursor):
    if row is None:
        return None
    columns = [desc[0] for desc in cursor.description]
    result = {}
    for col, val in zip(columns, row):
        if isinstance(val, (datetime, date)):
            result[col] = val.isoformat()
        else:
            result[col] = val
    return result


def rows_to_list(rows, cursor):
    columns = [desc[0] for desc in cursor.description]
    result = []
    for row in rows:
        d = {}
        for col, val in zip(columns, row):
            if isinstance(val, (datetime, date)):
                d[col] = val.isoformat()
            else:
                d[col] = val
        result.append(d)
    return result


def init_db():
    conn = get_db()
    c = conn.cursor()

    c.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            role TEXT DEFAULT 'user',
            department TEXT,
            avatar TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            active INTEGER DEFAULT 1
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS employees (
            id TEXT PRIMARY KEY,
            name TEXT NOT NULL,
            email TEXT,
            department TEXT,
            position TEXT,
            manager TEXT,
            admission_date TEXT,
            birth_date TEXT,
            phone TEXT,
            cpf TEXT,
            status TEXT DEFAULT 'active',
            photo TEXT,
            notes TEXT,
            exit_date TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS records (
            id TEXT PRIMARY KEY,
            area TEXT NOT NULL,
            module TEXT NOT NULL,
            title TEXT NOT NULL,
            content TEXT,
            employee_id TEXT,
            status TEXT DEFAULT 'active',
            priority TEXT DEFAULT 'medium',
            due_date TEXT,
            tags TEXT,
            attachments TEXT,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS performance_reviews (
            id TEXT PRIMARY KEY,
            employee_id TEXT NOT NULL,
            reviewer_id TEXT,
            period TEXT,
            scores TEXT,
            overall_score REAL,
            comments TEXT,
            status TEXT DEFAULT 'draft',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS surveys (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            description TEXT,
            questions TEXT,
            responses TEXT,
            status TEXT DEFAULT 'draft',
            start_date TEXT,
            end_date TEXT,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS events (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            description TEXT,
            event_type TEXT,
            date TEXT,
            time TEXT,
            location TEXT,
            attendees TEXT,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS news (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            content TEXT,
            category TEXT,
            author TEXT,
            published INTEGER DEFAULT 0,
            pinned INTEGER DEFAULT 0,
            image TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS org_chart (
            id TEXT PRIMARY KEY,
            employee_id TEXT NOT NULL,
            parent_id TEXT,
            position TEXT,
            level INTEGER DEFAULT 0,
            order_index INTEGER DEFAULT 0
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS knowledge_base (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            content TEXT,
            category TEXT,
            tags TEXT,
            author TEXT,
            views INTEGER DEFAULT 0,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS benefits (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            description TEXT,
            category TEXT,
            supplier TEXT,
            contact TEXT,
            url TEXT,
            active INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS recognitions (
            id TEXT PRIMARY KEY,
            employee_id TEXT NOT NULL,
            recognized_by TEXT,
            type TEXT,
            message TEXT,
            points INTEGER DEFAULT 0,
            public INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS service_requests (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            description TEXT,
            category TEXT,
            priority TEXT DEFAULT 'medium',
            status TEXT DEFAULT 'open',
            requester_id TEXT,
            assigned_to TEXT,
            resolution TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    c.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id TEXT PRIMARY KEY,
            user_id TEXT,
            action TEXT,
            resource TEXT,
            resource_id TEXT,
            details TEXT,
            ip_address TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        
        CREATE TABLE IF NOT EXISTS password_resets (
            id TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            email TEXT NOT NULL,
            token_hash TEXT NOT NULL,
            expires_at TIMESTAMP NOT NULL,
            used INTEGER DEFAULT 0,
            used_at TIMESTAMP,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS policies (
            id TEXT PRIMARY KEY,
            title TEXT NOT NULL,
            category TEXT NOT NULL,
            content TEXT,
            pdf_filename TEXT,
            created_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP
        )
    """)

    # Integração SSO com o Valore Hub — vínculo por hub_user_id.
    c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS hub_user_id TEXT")
    c.execute("DO $$ BEGIN "
              "IF NOT EXISTS (SELECT 1 FROM pg_constraint WHERE conname = 'users_hub_user_id_key') THEN "
              "ALTER TABLE users ADD CONSTRAINT users_hub_user_id_key UNIQUE (hub_user_id); "
              "END IF; END $$;")
    c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS hub_portal_role TEXT")
    c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS hub_session_version INTEGER")
    c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS last_sso_login TIMESTAMP")
    c.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS permission_checked_at TIMESTAMP")

    # Segmentação por departamento em notícias e eventos — NULL ou 'all' = todos.
    c.execute("ALTER TABLE news ADD COLUMN IF NOT EXISTS department TEXT")
    c.execute("ALTER TABLE events ADD COLUMN IF NOT EXISTS department TEXT")

    # Dono do registro (quem enviou) — usado para restringir Colaborador a ver
    # somente as próprias solicitações em Formulários (Operações).
    c.execute("ALTER TABLE records ADD COLUMN IF NOT EXISTS created_by_user_id TEXT")

    # Data de saída do colaborador, capturada na inativação.
    c.execute("ALTER TABLE employees ADD COLUMN IF NOT EXISTS exit_date TEXT")

    # Simplificação de perfis: 'rh' e 'gestor' deixam de existir — quem tinha
    # esses perfis passa a ser 'admin' (perfil único de acesso total). Idempotente.
    c.execute("UPDATE users SET role = 'admin' WHERE role IN ('rh', 'gestor')")

    # Seed admin user
    c.execute("SELECT id FROM users WHERE email = 'admin@empresa.com'")
    if not c.fetchone():
        hashed = bcrypt.hashpw('admin123'.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        c.execute("""INSERT INTO users (id, name, email, password_hash, role, department)
                     VALUES (%s, %s, %s, %s, %s, %s)""",
                  (str(uuid.uuid4()), 'Administrador', 'admin@empresa.com', hashed, 'admin', 'RH'))

        demo_employees = [
            ('Ana Silva', 'ana.silva@empresa.com', 'RH', 'Gerente de RH', '2021-03-15', '1988-07-22'),
            ('Carlos Mendes', 'carlos.mendes@empresa.com', 'Tecnologia', 'Desenvolvedor Sênior', '2020-06-01', '1990-11-05'),
            ('Fernanda Lima', 'fernanda.lima@empresa.com', 'Financeiro', 'Analista Financeiro', '2022-01-10', '1993-04-18'),
            ('Roberto Costa', 'roberto.costa@empresa.com', 'Comercial', 'Gerente Comercial', '2019-08-20', '1985-09-30'),
            ('Juliana Rocha', 'juliana.rocha@empresa.com', 'Marketing', 'Analista de Marketing', '2023-02-14', '1995-12-03'),
            ('Marcos Andrade', 'marcos.andrade@empresa.com', 'Operações', 'Coordenador Operacional', '2021-11-08', '1987-06-25'),
        ]
        for emp in demo_employees:
            c.execute("""INSERT INTO employees (id, name, email, department, position, admission_date, birth_date, status)
                         VALUES (%s, %s, %s, %s, %s, %s, %s, 'active')""",
                      (str(uuid.uuid4()), emp[0], emp[1], emp[2], emp[3], emp[4], emp[5]))

        news_items = [
            ('Bem-vindo ao novo Sistema de Gestão de Pessoas!', 'Estamos lançando nossa plataforma integrada...', 'Comunicado', 1, 1),
            ('Programa de Metas Q3 2026', 'Acesse os objetivos do trimestre...', 'Metas', 1, 0),
            ('Treinamento de Liderança - Vagas Abertas', 'Inscrições abertas para o programa de liderança...', 'Treinamento', 1, 0),
        ]
        for n in news_items:
            c.execute("""INSERT INTO news (id, title, content, category, author, published, pinned)
                         VALUES (%s, %s, %s, %s, %s, %s, %s)""",
                      (str(uuid.uuid4()), n[0], n[1], n[2], 'Administrador', n[3], n[4]))

        benefits_items = [
            ('Plano de Saúde', 'Cobertura médica e hospitalar completa', 'Saúde', 'Unimed'),
            ('Vale Refeição', 'R$ 35,00 por dia útil', 'Alimentação', 'Sodexo'),
            ('Gympass', 'Acesso a academias parceiras', 'Bem-estar', 'Gympass'),
            ('PLR', 'Participação nos Lucros e Resultados', 'Financeiro', 'Empresa'),
            ('Seguro de Vida', 'Cobertura para colaboradores e dependentes', 'Seguridade', 'Porto Seguro'),
        ]
        for b in benefits_items:
            c.execute("""INSERT INTO benefits (id, title, description, category, supplier, active)
                         VALUES (%s, %s, %s, %s, %s, 1)""",
                      (str(uuid.uuid4()), b[0], b[1], b[2], b[3]))

    conn.commit()
    conn.close()


def log_action(user_id, action, resource, resource_id='', details=''):
    try:
        conn = get_db()
        c = conn.cursor()
        c.execute("""INSERT INTO audit_log (id, user_id, action, resource, resource_id, details)
                     VALUES (%s, %s, %s, %s, %s, %s)""",
                  (str(uuid.uuid4()), user_id, action, resource, resource_id, details))
        conn.commit()
        conn.close()
    except:
        pass


def login_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return jsonify({'error': 'Não autorizado'}), 401
        return f(*args, **kwargs)
    return decorated


# ── SSO — Introspecção de permissão (Hub) ───────────────────────────────────
# Garante que revogações/alterações de acesso feitas no Hub (bloqueio,
# mudança de perfil, logout forçado) se refletem aqui em poucos segundos,
# em vez de só no próximo login.
SSO_PERMISSION_CACHE_SECONDS = 60


def checar_permissao_no_hub(hub_user_id):
    """Consulta o endpoint de introspecção do Hub. Propaga requests.RequestException em falha de rede."""
    resp = requests.post(
        f'{HUB_BASE_URL}/api/sso/introspect/',
        json={'client_id': SSO_CLIENT_ID, 'client_secret': SSO_CLIENT_SECRET, 'sub': hub_user_id},
        timeout=5,
    )
    resp.raise_for_status()
    return resp.json()


@app.before_request
def verificar_permissao_sso():
    """Revalida periodicamente, junto ao Hub, se o usuário autenticado ainda tem acesso.

    Só se aplica a usuários vinculados via SSO (hub_user_id preenchido) e usa um
    cache de SSO_PERMISSION_CACHE_SECONDS segundos para não bater no Hub a cada requisição.
    Falhas de rede ao consultar o Hub não bloqueiam a requisição (fail open).
    """
    if 'user_id' not in session:
        return None

    user_id = session['user_id']
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute(
            "SELECT hub_user_id, hub_session_version, permission_checked_at FROM users WHERE id = %s",
            (user_id,),
        )
        row = c.fetchone()
        if not row:
            return None
        hub_user_id, hub_session_version, permission_checked_at = row

        if not hub_user_id:
            return None  # conta local (pré-existente), sem vínculo SSO

        agora = datetime.utcnow()
        precisa_checar = (
            permission_checked_at is None or
            (agora - permission_checked_at).total_seconds() > SSO_PERMISSION_CACHE_SECONDS
        )
        if not precisa_checar:
            return None

        try:
            resultado = checar_permissao_no_hub(hub_user_id)
        except requests.RequestException as exc:
            app.logger.warning('Falha ao consultar introspecção do Hub para usuário %s: %s', user_id, exc)
            return None

        novo_session_version = resultado.get('session_version')
        sessao_invalidada = (
            not resultado.get('active', False) or
            (novo_session_version is not None and hub_session_version is not None and
             novo_session_version > hub_session_version)
        )
        if sessao_invalidada:
            session.clear()
            if request.path.startswith('/api/'):
                return jsonify({'error': 'Sua sessão expirou ou seu acesso foi revogado no Valore Hub.'}), 401
            return redirect(f'{HUB_BASE_URL}/login/')

        c.execute(
            "UPDATE users SET hub_portal_role = %s, hub_session_version = %s, permission_checked_at = %s WHERE id = %s",
            (resultado.get('portal_role'), novo_session_version, agora, user_id),
        )
        conn.commit()
    finally:
        conn.close()
    return None


# ── AUTH ──────────────────────────────────────────────────────────────────────
@app.route('/api/auth/login', methods=['POST'])
def login():
    data = request.get_json()
    email = data.get('email', '').strip().lower()
    password = data.get('password', '')

    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM users WHERE email = %s AND active = 1", (email,))
    row = c.fetchone()
    user = row_to_dict(row, c) if row else None
    conn.close()

    if not user:
        return jsonify({'error': 'Credenciais inválidas'}), 401
    if not bcrypt.checkpw(password.encode('utf-8'), user['password_hash'].encode('utf-8')):
        return jsonify({'error': 'Credenciais inválidas'}), 401

    session['user_id'] = user['id']
    session['user_name'] = user['name']
    session['user_role'] = user['role']
    session.permanent = True

    log_action(user['id'], 'LOGIN', 'auth')
    return jsonify({'user': {'id': user['id'], 'name': user['name'], 'email': user['email'],
                             'role': user['role'], 'department': user['department']}})


@app.route('/api/auth/logout', methods=['GET', 'POST'])
def logout():
    from flask import redirect
    session.clear()
    return redirect('/')


@app.route('/auth/sso/callback', methods=['GET'])
def sso_callback():
    code = request.args.get('code', '').strip()
    if not code:
        return redirect('/?erro=sso_sem_codigo')

    try:
        claims = trocar_codigo_por_claims(code)
        user = localizar_ou_criar_usuario_via_sso(claims)
    except SSOClientError as exc:
        app.logger.warning('Falha no login SSO: %s', exc)
        return redirect('/?erro=sso_falhou')

    session['user_id'] = user['id']
    session['user_name'] = user['name']
    session['user_role'] = user['role']
    session.permanent = True

    log_action(user['id'], 'LOGIN', 'auth', details='via Valore Hub (SSO)')
    return redirect(SSO_DASHBOARD_PATH)


def send_reset_email(to_email, reset_link, user_name):
    gmail_user = os.environ.get('GMAIL_USER')
    gmail_password = os.environ.get('GMAIL_PASSWORD')

    if not gmail_user or not gmail_password:
        raise ValueError('GMAIL_USER e GMAIL_PASSWORD nao configurados no .env')

    subject = 'Redefinicao de Senha — Valore RH'

    html_body = (
        '<div style="font-family: Arial, sans-serif; background: #f4f4f4; margin: 0; padding: 20px;">'
        '<div style="max-width: 520px; margin: 0 auto; background: #ffffff; border-radius: 8px; overflow: hidden;">'
        '<div style="background: #1d4ed8; padding: 28px 32px;">'
        '<h1 style="color: #ffffff; margin: 0; font-size: 20px;">Valore RH</h1>'
        '<p style="color: #bfdbfe; margin: 4px 0 0;">Redefinicao de Senha</p>'
        '</div>'
        '<div style="padding: 32px;">'
        '<p style="color: #374151; font-size: 15px;">Ola, <strong>' + user_name + '</strong>!</p>'
        '<p style="color: #374151; font-size: 15px;">'
        'Recebemos uma solicitacao para redefinir a senha da sua conta no Valore RH. '
        'Clique no botao abaixo para criar uma nova senha.'
        '</p>'
        '<div style="text-align: center; margin: 32px 0;">'
        '<a href="' + reset_link + '" style="background: #1d4ed8; color: #ffffff; padding: 14px 28px; '
        'border-radius: 6px; text-decoration: none; font-weight: bold; font-size: 15px; display: inline-block;">'
        'Redefinir Minha Senha'
        '</a>'
        '</div>'
        '<p style="color: #6b7280; font-size: 13px;">'
        'Este link expira em <strong>1 hora</strong>.'
        '</p>'
        '<p style="color: #6b7280; font-size: 13px;">'
        'Se voce nao solicitou isso, pode ignorar este e-mail com seguranca. '
        'Sua senha nao sera alterada.'
        '</p>'
        '<hr style="border: none; border-top: 1px solid #e5e7eb; margin: 24px 0;">'
        '<p style="color: #9ca3af; font-size: 12px;">'
        'Se o botao nao funcionar, copie e cole este link no navegador:<br>'
        '<a href="' + reset_link + '" style="color: #1d4ed8;">' + reset_link + '</a>'
        '</p>'
        '</div>'
        '<div style="background: #f9fafb; padding: 16px 32px; text-align: center;">'
        '<p style="color: #9ca3af; font-size: 12px; margin: 0;">'
        'Este e um e-mail automatico. Por favor, nao responda.'
        '</p>'
        '</div>'
        '</div>'
        '</div>'
    )

    msg = MIMEMultipart('alternative')
    msg['Subject'] = subject
    msg['From'] = 'Valore RH <' + gmail_user + '>'
    msg['To'] = to_email
    msg.attach(MIMEText(html_body, 'html', 'utf-8'))

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(gmail_user, gmail_password)
        server.sendmail(gmail_user, to_email, msg.as_string())


# ── Notificações por e-mail — Notícias e Eventos ─────────────────────────────
def departamento_visivel_para(department_value, user_department):
    """Verifica se um valor de `department` (NULL, 'all' ou lista separada por
    vírgula) é visível para o departamento do usuário logado."""
    if not department_value or department_value == 'all':
        return True
    alvos = [d.strip() for d in department_value.split(',') if d.strip()]
    return user_department in alvos


def obter_departamento_usuario(user_id):
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("SELECT department FROM users WHERE id = %s", (user_id,))
        row = c.fetchone()
        return row[0] if row else None
    finally:
        conn.close()


def emails_funcionarios_por_departamento(department_value):
    """Retorna os e-mails de funcionários ativos alvo de uma notícia/evento."""
    conn = get_db()
    c = conn.cursor()
    try:
        if not department_value or department_value == 'all':
            c.execute("SELECT email FROM employees WHERE status = 'active' AND email IS NOT NULL AND email != ''")
        else:
            alvos = [d.strip() for d in department_value.split(',') if d.strip()]
            if not alvos:
                c.execute("SELECT email FROM employees WHERE status = 'active' AND email IS NOT NULL AND email != ''")
            else:
                c.execute(
                    "SELECT email FROM employees WHERE status = 'active' AND email IS NOT NULL "
                    "AND email != '' AND department IN %s",
                    (tuple(alvos),)
                )
        return [row[0] for row in c.fetchall() if row[0] and '@' in row[0]]
    finally:
        conn.close()


def enviar_notificacao_em_massa(subject, html_body, to_emails):
    gmail_user = os.environ.get('GMAIL_USER')
    gmail_password = os.environ.get('GMAIL_PASSWORD')

    if not gmail_user or not gmail_password:
        raise ValueError('GMAIL_USER e GMAIL_PASSWORD nao configurados no .env')

    with smtplib.SMTP_SSL('smtp.gmail.com', 465) as server:
        server.login(gmail_user, gmail_password)
        for to_email in to_emails:
            try:
                msg = MIMEMultipart('alternative')
                msg['Subject'] = subject
                msg['From'] = 'Valore RH <' + gmail_user + '>'
                msg['To'] = to_email
                msg.attach(MIMEText(html_body, 'html', 'utf-8'))
                server.sendmail(gmail_user, to_email, msg.as_string())
            except Exception as exc:
                # Um endereço inválido não pode travar o envio para o resto
                # da lista (ex.: e-mail cadastrado errado em um funcionário).
                app.logger.warning('Falha ao enviar notificação para %s: %s', to_email, exc)


def montar_email_notificacao(titulo_topo, subtitulo, linhas_html):
    app_url = os.environ.get('APP_URL', 'https://rh.valore.com.br')
    return (
        '<div style="font-family: Arial, sans-serif; background: #f4f4f4; margin: 0; padding: 20px;">'
        '<div style="max-width: 520px; margin: 0 auto; background: #ffffff; border-radius: 8px; overflow: hidden;">'
        '<div style="background: #1d4ed8; padding: 28px 32px;">'
        '<h1 style="color: #ffffff; margin: 0; font-size: 20px;">Valore RH</h1>'
        '<p style="color: #bfdbfe; margin: 4px 0 0;">' + subtitulo + '</p>'
        '</div>'
        '<div style="padding: 32px;">'
        + linhas_html +
        '<div style="text-align: center; margin: 32px 0;">'
        '<a href="' + app_url + '" style="background: #1d4ed8; color: #ffffff; padding: 14px 28px; '
        'border-radius: 6px; text-decoration: none; font-weight: bold; font-size: 15px; display: inline-block;">'
        'Acessar o Valore RH'
        '</a>'
        '</div>'
        '</div>'
        '<div style="background: #f9fafb; padding: 16px 32px; text-align: center;">'
        '<p style="color: #9ca3af; font-size: 12px; margin: 0;">'
        'Este e um e-mail automatico. Por favor, nao responda.'
        '</p>'
        '</div>'
        '</div>'
        '</div>'
    )


@app.route('/api/auth/forgot-password', methods=['POST'])
def forgot_password():
    data = request.get_json()
    email = data.get('email', '').strip().lower()

    if not email:
        return jsonify({'error': 'E-mail e obrigatorio'}), 400

    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT id, name, email FROM users WHERE email = %s AND active = 1', (email,))
    user = c.fetchone()

    if not user:
        c.close()
        conn.close()
        return jsonify({'message': 'Se este e-mail estiver cadastrado, voce recebera as instrucoes em breve.'}), 200

    user_id = user[0]
    user_name = user[1]
    user_email = user[2]

    c.execute('UPDATE password_resets SET used = 1 WHERE user_id = %s AND used = 0', (user_id,))

    raw_token = secrets.token_urlsafe(48)
    token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
    expires_at = datetime.utcnow() + timedelta(hours=1)
    reset_id = str(uuid.uuid4())

    c.execute(
        'INSERT INTO password_resets (id, user_id, email, token_hash, expires_at) VALUES (%s, %s, %s, %s, %s)',
        (reset_id, user_id, user_email, token_hash, expires_at)
    )
    conn.commit()
    c.close()
    conn.close()

    app_url = os.environ.get('APP_URL', 'https://rh.valore.com.br')
    reset_link = app_url + '/reset-senha?token=' + raw_token

    try:
        send_reset_email(user_email, reset_link, user_name)
    except Exception as e:
        print('[ERRO EMAIL] ' + str(e))
        return jsonify({'error': 'Falha ao enviar e-mail. Contate o administrador.'}), 500

    return jsonify({'message': 'Se este e-mail estiver cadastrado, voce recebera as instrucoes em breve.'}), 200


@app.route('/api/auth/reset-password', methods=['POST'])
def reset_password():
    data = request.get_json()
    raw_token = data.get('token', '').strip()
    new_password = data.get('password', '')

    if not raw_token or not new_password:
        return jsonify({'error': 'Token e nova senha sao obrigatorios'}), 400

    if len(new_password) < 8:
        return jsonify({'error': 'A senha deve ter pelo menos 8 caracteres'}), 400

    token_hash = hashlib.sha256(raw_token.encode()).hexdigest()
    now = datetime.utcnow()

    conn = get_db()
    c = conn.cursor()
    c.execute(
        'SELECT * FROM password_resets WHERE token_hash = %s AND used = 0 AND expires_at > %s',
        (token_hash, now)
    )
    reset = c.fetchone()

    if not reset:
        c.close()
        conn.close()
        return jsonify({'error': 'Link invalido ou expirado. Solicite um novo.'}), 400

    user_id = reset[1]
    hashed = bcrypt.hashpw(new_password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    c.execute('UPDATE users SET password_hash = %s WHERE id = %s', (hashed, user_id))
    c.execute('UPDATE password_resets SET used = 1, used_at = %s WHERE id = %s', (now, reset[0]))
    conn.commit()
    c.close()
    conn.close()

    return jsonify({'message': 'Senha redefinida com sucesso! Voce ja pode fazer login.'}), 200




@app.route('/api/auth/me', methods=['GET'])
@login_required
def me():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, name, email, role, department FROM users WHERE id = %s", (session['user_id'],))
    row = c.fetchone()
    user = row_to_dict(row, c) if row else None
    conn.close()
    if user:
        return jsonify(user)
    return jsonify({'error': 'Usuário não encontrado'}), 404


# ── DASHBOARD ─────────────────────────────────────────────────────────────────
@app.route('/api/dashboard/stats', methods=['GET'])
@login_required
def dashboard_stats():
    conn = get_db()
    c = conn.cursor()

    c.execute("SELECT COUNT(*) FROM employees WHERE status='active'")
    employees = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM news WHERE published=1")
    news_count = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM service_requests WHERE status='open'")
    open_requests = c.fetchone()[0]

    c.execute("SELECT COUNT(*) FROM recognitions")
    recognitions = c.fetchone()[0]

    month = date.today().month
    c.execute("""SELECT name, birth_date FROM employees
                 WHERE status='active' AND birth_date IS NOT NULL
                 AND birth_date != ''
                 AND EXTRACT(MONTH FROM birth_date::date) = %s""", (month,))
    birthdays = rows_to_list(c.fetchall(), c)

    c.execute("""SELECT name, admission_date FROM employees
                 WHERE status='active' AND admission_date IS NOT NULL
                 AND admission_date != ''
                 AND EXTRACT(MONTH FROM admission_date::date) = %s
                 AND EXTRACT(YEAR FROM admission_date::date) < %s""", (month, date.today().year))
    work_anniversaries = rows_to_list(c.fetchall(), c)

    c.execute("SELECT id, title, category, created_at FROM news WHERE published=1 ORDER BY created_at DESC LIMIT 5")
    recent_news = rows_to_list(c.fetchall(), c)

    c.execute("SELECT department, COUNT(*) as c FROM employees WHERE status='active' GROUP BY department ORDER BY c DESC")
    dept_counts = rows_to_list(c.fetchall(), c)

    conn.close()
    return jsonify({
        'employees': employees,
        'news': news_count,
        'open_requests': open_requests,
        'recognitions': recognitions,
        'birthdays_this_month': birthdays,
        'work_anniversaries_this_month': work_anniversaries,
        'recent_news': recent_news,
        'dept_counts': dept_counts
    })


# ── EMPLOYEES ─────────────────────────────────────────────────────────────────
@app.route('/api/employees', methods=['GET'])
@login_required
def get_employees():
    search = request.args.get('search', '')
    dept = request.args.get('department', '')
    status = request.args.get('status', 'active')
    conn = get_db()
    c = conn.cursor()
    query = "SELECT * FROM employees WHERE 1=1"
    params = []
    if status:
        query += " AND status = %s"
        params.append(status)
    if search:
        query += " AND (name ILIKE %s OR email ILIKE %s OR position ILIKE %s)"
        params.extend([f'%{search}%'] * 3)
    if dept:
        query += " AND department = %s"
        params.append(dept)
    query += " ORDER BY name ASC"
    c.execute(query, params)
    employees = rows_to_list(c.fetchall(), c)
    conn.close()
    return jsonify(employees)


@app.route('/api/employees', methods=['POST'])
@login_required
def create_employee():
    data = request.get_json()
    emp_id = str(uuid.uuid4())
    conn = get_db()
    c = conn.cursor()
    c.execute("""INSERT INTO employees (id, name, email, department, position, manager,
                 admission_date, birth_date, phone, cpf, status, notes, is_parent, parent_type)
                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
              (emp_id, data.get('name'), data.get('email'), data.get('department'),
               data.get('position'), data.get('manager'), data.get('admission_date'),
               data.get('birth_date'), data.get('phone'), data.get('cpf'),
               data.get('status', 'active'), data.get('notes'),
               data.get('is_parent', False), data.get('parent_type', '')))
    conn.commit()
    conn.close()
    log_action(session['user_id'], 'CREATE', 'employee', emp_id, data.get('name'))
    provisionar_usuario_hub(data.get('name'), data.get('email'), data.get('department'), data.get('position'))
    return jsonify({'id': emp_id, 'success': True}), 201


@app.route('/api/employees/<emp_id>', methods=['GET'])
@login_required
def get_employee(emp_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM employees WHERE id = %s", (emp_id,))
    row = c.fetchone()
    emp = row_to_dict(row, c) if row else None
    conn.close()
    if not emp:
        return jsonify({'error': 'Não encontrado'}), 404
    return jsonify(emp)


@app.route('/api/employees/<emp_id>', methods=['PUT'])
@login_required
def update_employee(emp_id):
    data = request.get_json()
    conn = get_db()
    c = conn.cursor()
    c.execute("""UPDATE employees SET name=%s, email=%s, department=%s, position=%s, manager=%s,
                 admission_date=%s, birth_date=%s, phone=%s, cpf=%s, status=%s, notes=%s,
                 is_parent=%s, parent_type=%s, exit_date=%s, updated_at=CURRENT_TIMESTAMP WHERE id=%s""",
              (data.get('name'), data.get('email'), data.get('department'), data.get('position'),
               data.get('manager'), data.get('admission_date'), data.get('birth_date'),
               data.get('phone'), data.get('cpf'), data.get('status'), data.get('notes'),
               data.get('is_parent', False), data.get('parent_type', ''),
               data.get('exit_date') or None, emp_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/employees/<emp_id>', methods=['DELETE'])
@login_required
def delete_employee(emp_id):
    data = request.get_json(silent=True) or {}
    exit_date = data.get('exit_date')
    reason = (data.get('notes') or '').strip()

    conn = get_db()
    c = conn.cursor()

    notes_suffix = f"Motivo da saída: {reason}" if reason else None
    if notes_suffix:
        c.execute("""UPDATE employees SET status='inactive', exit_date=%s,
                     notes = CASE WHEN notes IS NULL OR notes = '' THEN %s
                                  ELSE notes || %s END,
                     updated_at=CURRENT_TIMESTAMP WHERE id=%s""",
                  (exit_date, notes_suffix, '\n' + notes_suffix, emp_id))
    else:
        c.execute("""UPDATE employees SET status='inactive', exit_date=%s,
                     updated_at=CURRENT_TIMESTAMP WHERE id=%s""",
                  (exit_date, emp_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ── RECORDS ───────────────────────────────────────────────────────────────────
@app.route('/api/records', methods=['GET'])
@login_required
def get_records():
    area = request.args.get('area', '')
    module = request.args.get('module', '')
    employee_id = request.args.get('employee_id', '')
    search = request.args.get('search', '')
    conn = get_db()
    c = conn.cursor()
    query = "SELECT * FROM records WHERE 1=1"
    params = []
    if area:
        query += " AND area = %s"
        params.append(area)
    if module:
        query += " AND module = %s"
        params.append(module)
    if employee_id:
        query += " AND employee_id = %s"
        params.append(employee_id)
    if search:
        query += " AND (title ILIKE %s OR content ILIKE %s)"
        params.extend([f'%{search}%', f'%{search}%'])
    # Colaborador só pode ver as próprias solicitações em Formulários
    # (Operações) — filtro forçado no servidor (não confia em parâmetro vindo
    # do cliente). Liderança e Administrador continuam vendo todas, pois
    # precisam aprovar/recusar. Escopo limitado a este módulo para não afetar
    # a visibilidade de Colaborador em outras páginas genéricas (Férias,
    # Documentos Admissionais etc.), que hoje listam registros de terceiros.
    if session.get('user_role') == 'colaborador' and area == 'operacoes' and module == 'formularios':
        query += " AND created_by_user_id = %s"
        params.append(session['user_id'])
    query += " ORDER BY created_at DESC"
    c.execute(query, params)
    records = rows_to_list(c.fetchall(), c)
    conn.close()
    return jsonify(records)


def _pode_editar_record(area, module):
    """PDI (área dev_talentos / módulo pdv) e Formulários (área operacoes /
    módulo formularios) podem ser editados/aprovados por Liderança além de
    Administrador; todas as demais áreas (Aprendizado, Documentos
    Admissionais, Férias, Cultura, etc.) são restritas a Administrador."""
    role = session.get('user_role')
    if role == 'admin':
        return True
    if role == 'lideranca' and area == 'dev_talentos' and module == 'pdv':
        return True
    if role == 'lideranca' and area == 'operacoes' and module == 'formularios':
        return True
    return False


def _pode_criar_record(area, module):
    """Quem pode enviar (POST) um novo registro. Quem pode editar sempre pode
    criar; além disso, Formulários (Operações) aceita envio de QUALQUER
    usuário autenticado (Colaborador, Liderança, Administrador) — é a própria
    tela de solicitação. A aprovação/mudança de status continua restrita a
    Liderança/Administrador via _pode_editar_record."""
    if _pode_editar_record(area, module):
        return True
    if area == 'operacoes' and module == 'formularios':
        return True
    return False


@app.route('/api/records', methods=['POST'])
@login_required
def create_record():
    data = request.get_json()
    area = data.get('area')
    module = data.get('module')
    if not _pode_criar_record(area, module):
        return jsonify({'error': 'Acesso negado'}), 403
    rec_id = str(uuid.uuid4())
    content = json.dumps(data.get('content', {})) if isinstance(data.get('content'), dict) else data.get('content', '')
    tags = json.dumps(data.get('tags', [])) if isinstance(data.get('tags'), list) else data.get('tags', '')
    # Formulários (Operações) usa 'pendente'/'aprovado'/'recusado' como status
    # padrão (em vez do genérico 'active' usado pelas demais áreas).
    default_status = 'pendente' if module == 'formularios' else 'active'
    conn = get_db()
    c = conn.cursor()
    c.execute("""INSERT INTO records (id, area, module, title, content, employee_id, status, priority, due_date, tags, created_by, created_by_user_id)
                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)""",
              (rec_id, area, module, data.get('title'), content,
               data.get('employee_id'), data.get('status', default_status), data.get('priority', 'medium'),
               data.get('due_date'), tags, session['user_id'], session['user_id']))
    conn.commit()
    conn.close()
    log_action(session['user_id'], 'CREATE', f"{area}/{module}", rec_id, data.get('title'))
    return jsonify({'id': rec_id, 'success': True}), 201


@app.route('/api/records/<rec_id>', methods=['PUT'])
@login_required
def update_record(rec_id):
    data = request.get_json()
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT area, module FROM records WHERE id = %s", (rec_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Não encontrado'}), 404
    if not _pode_editar_record(row[0], row[1]):
        conn.close()
        return jsonify({'error': 'Acesso negado'}), 403
    content = json.dumps(data.get('content', {})) if isinstance(data.get('content'), dict) else data.get('content', '')
    c.execute("""UPDATE records SET title=%s, content=%s, status=%s, priority=%s, due_date=%s,
                 updated_at=CURRENT_TIMESTAMP WHERE id=%s""",
              (data.get('title'), content, data.get('status'), data.get('priority'), data.get('due_date'), rec_id))
    conn.commit()

    # Aprovação de Solicitação de Férias/Licença (Formulários/Operações) cria
    # automaticamente um registro espelho em rh/ferias com o período
    # preenchido, para aparecer na página de Férias. Feito aqui no PUT
    # (server-side, não depende do cliente). Uma falha aqui NÃO deve derrubar
    # a resposta de sucesso da troca de status — só loga um aviso.
    if row[0] == 'operacoes' and row[1] == 'formularios' and data.get('status') == 'aprovado':
        try:
            c.execute("SELECT title, content FROM records WHERE id = %s", (rec_id,))
            rec_row = c.fetchone()
            if rec_row and rec_row[0] in ('Solicitação de Férias', 'Solicitação de Licença'):
                rec_content = json.loads(rec_row[1]) if rec_row[1] else {}
                data_inicio = rec_content.get('data_inicio', '')
                data_fim = rec_content.get('data_fim', '')
                if data_inicio and data_fim:
                    tipo = 'ferias' if rec_row[0] == 'Solicitação de Férias' else 'licenca_medica'
                    periodo = f"{data_inicio} a {data_fim}"
                    ferias_title = rec_row[0] + (' — ' + session.get('user_name', '') if session.get('user_name') else '')
                    ferias_content = {'type': tipo, 'period': periodo}
                    c.execute(
                        """INSERT INTO records (id, area, module, title, content, status, priority, created_by, created_by_user_id)
                           VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
                        (str(uuid.uuid4()), 'rh', 'ferias', ferias_title, json.dumps(ferias_content),
                         'active', 'medium', session['user_id'], session['user_id']))
                    conn.commit()
        except Exception as exc:
            app.logger.warning('Falha ao criar registro de férias a partir da aprovação do formulário %s: %s', rec_id, exc)

    conn.close()
    return jsonify({'success': True})


@app.route('/api/records/<rec_id>', methods=['DELETE'])
@login_required
def delete_record(rec_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT area, module FROM records WHERE id = %s", (rec_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({'error': 'Não encontrado'}), 404
    if not _pode_editar_record(row[0], row[1]):
        conn.close()
        return jsonify({'error': 'Acesso negado'}), 403
    c.execute("DELETE FROM records WHERE id=%s", (rec_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ── DEPARTAMENTOS ─────────────────────────────────────────────────────────────
@app.route('/api/departments', methods=['GET'])
@login_required
def get_departments():
    conn = get_db()
    c = conn.cursor()
    c.execute("""SELECT DISTINCT department FROM employees
                 WHERE status='active' AND department IS NOT NULL AND department != ''
                 ORDER BY department""")
    departments = [row[0] for row in c.fetchall()]
    conn.close()
    return jsonify(departments)


# ── NEWS ──────────────────────────────────────────────────────────────────────
@app.route('/api/news', methods=['GET'])
@login_required
def get_news():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM news ORDER BY pinned DESC, created_at DESC")
    news = rows_to_list(c.fetchall(), c)
    conn.close()
    user_department = obter_departamento_usuario(session['user_id'])
    news = [n for n in news if departamento_visivel_para(n.get('department'), user_department)]
    return jsonify(news)


@app.route('/api/news', methods=['POST'])
@login_required
def create_news():
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    data = request.get_json()
    news_id = str(uuid.uuid4())
    department = data.get('department') or 'all'
    conn = get_db()
    c = conn.cursor()
    c.execute("""INSERT INTO news (id, title, content, category, author, published, pinned, department)
                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
              (news_id, data.get('title'), data.get('content'), data.get('category'),
               session.get('user_name', 'Sistema'), data.get('published', 0), data.get('pinned', 0), department))
    conn.commit()
    conn.close()

    try:
        recipients = emails_funcionarios_por_departamento(department)
        if recipients:
            excerpt = (data.get('content') or '')[:200]
            corpo = montar_email_notificacao(
                'Nova Notícia',
                'Nova Notícia',
                '<p style="color: #374151; font-size: 15px;">Uma nova notícia foi publicada no Valore RH:</p>'
                '<p style="color: #111827; font-size: 16px; font-weight: bold; margin: 16px 0 6px;">' + (data.get('title') or '') + '</p>'
                '<p style="color: #6b7280; font-size: 13px;">' + excerpt + ('...' if len(data.get('content') or '') > 200 else '') + '</p>'
                '<p style="color: #374151; font-size: 14px;">Acesse o portal para ler a notícia completa.</p>'
            )
            enviar_notificacao_em_massa('Nova notícia: ' + (data.get('title') or ''), corpo, recipients)
    except Exception as exc:
        app.logger.warning('Falha ao enviar e-mails de notificação de notícia %s: %s', news_id, exc)

    return jsonify({'id': news_id, 'success': True}), 201


@app.route('/api/news/<news_id>', methods=['PUT'])
@login_required
def update_news(news_id):
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    data = request.get_json()
    conn = get_db()
    c = conn.cursor()
    c.execute("""UPDATE news SET title=%s, content=%s, category=%s, published=%s, pinned=%s,
                 updated_at=CURRENT_TIMESTAMP WHERE id=%s""",
              (data.get('title'), data.get('content'), data.get('category'),
               data.get('published', 0), data.get('pinned', 0), news_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/news/<news_id>', methods=['DELETE'])
@login_required
def delete_news(news_id):
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM news WHERE id=%s", (news_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ── EVENTS ────────────────────────────────────────────────────────────────────
@app.route('/api/events', methods=['GET'])
@login_required
def get_events():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM events ORDER BY date ASC")
    events = rows_to_list(c.fetchall(), c)
    conn.close()
    user_department = obter_departamento_usuario(session['user_id'])
    events = [e for e in events if departamento_visivel_para(e.get('department'), user_department)]
    return jsonify(events)


@app.route('/api/events', methods=['POST'])
@login_required
def create_event():
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    data = request.get_json()
    ev_id = str(uuid.uuid4())
    department = data.get('department') or 'all'
    conn = get_db()
    c = conn.cursor()
    c.execute("""INSERT INTO events (id, title, description, event_type, date, time, location, created_by, department)
                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)""",
              (ev_id, data.get('title'), data.get('description'), data.get('event_type'),
               data.get('date'), data.get('time'), data.get('location'), session['user_id'], department))
    conn.commit()
    conn.close()

    try:
        recipients = emails_funcionarios_por_departamento(department)
        if recipients:
            corpo = montar_email_notificacao(
                'Novo Evento',
                'Novo Evento',
                '<p style="color: #374151; font-size: 15px;">Um novo evento foi agendado no Valore RH:</p>'
                '<p style="color: #111827; font-size: 16px; font-weight: bold; margin: 16px 0 6px;">' + (data.get('title') or '') + '</p>'
                '<p style="color: #374151; font-size: 14px;">Data: ' + (data.get('date') or '—') + '</p>'
                '<p style="color: #374151; font-size: 14px;">Horário: ' + (data.get('time') or '—') + '</p>'
                '<p style="color: #374151; font-size: 14px;">Local: ' + (data.get('location') or '—') + '</p>'
                '<p style="color: #6b7280; font-size: 13px;">' + (data.get('description') or '') + '</p>'
            )
            enviar_notificacao_em_massa('Novo evento: ' + (data.get('title') or ''), corpo, recipients)
    except Exception as exc:
        app.logger.warning('Falha ao enviar e-mails de notificação de evento %s: %s', ev_id, exc)

    return jsonify({'id': ev_id, 'success': True}), 201


# ── BENEFITS ──────────────────────────────────────────────────────────────────
@app.route('/api/benefits', methods=['GET'])
@login_required
def get_benefits():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM benefits WHERE active=1 ORDER BY category, title")
    benefits = rows_to_list(c.fetchall(), c)
    conn.close()
    return jsonify(benefits)


@app.route('/api/benefits', methods=['POST'])
@login_required
def create_benefit():
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    data = request.get_json()
    b_id = str(uuid.uuid4())
    conn = get_db()
    c = conn.cursor()
    c.execute("""INSERT INTO benefits (id, title, description, category, supplier, contact, url)
                 VALUES (%s, %s, %s, %s, %s, %s, %s)""",
              (b_id, data.get('title'), data.get('description'), data.get('category'),
               data.get('supplier'), data.get('contact'), data.get('url')))
    conn.commit()
    conn.close()
    return jsonify({'id': b_id, 'success': True}), 201


@app.route('/api/benefits/<b_id>', methods=['DELETE'])
@login_required
def delete_benefit(b_id):
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE benefits SET active=0 WHERE id=%s", (b_id,))
    conn.commit()
    conn.close()
    log_action(session['user_id'], 'DELETE', 'benefit', b_id)
    return jsonify({'success': True})


# ── RECOGNITIONS ──────────────────────────────────────────────────────────────
@app.route('/api/recognitions', methods=['GET'])
@login_required
def get_recognitions():
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT r.*, e.name as employee_name, e.department, e.position
        FROM recognitions r
        LEFT JOIN employees e ON r.employee_id = e.id
        ORDER BY r.created_at DESC LIMIT 50
    """)
    recs = rows_to_list(c.fetchall(), c)
    conn.close()
    return jsonify(recs)


@app.route('/api/recognitions', methods=['POST'])
@login_required
def create_recognition():
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    data = request.get_json()
    r_id = str(uuid.uuid4())
    conn = get_db()
    c = conn.cursor()
    c.execute("""INSERT INTO recognitions (id, employee_id, recognized_by, type, message, points, public)
                 VALUES (%s, %s, %s, %s, %s, %s, %s)""",
              (r_id, data.get('employee_id'), session.get('user_name', 'Sistema'),
               data.get('type'), data.get('message'), data.get('points', 0), data.get('public', 1)))
    conn.commit()
    conn.close()
    return jsonify({'id': r_id, 'success': True}), 201


# ── KNOWLEDGE BASE ────────────────────────────────────────────────────────────
@app.route('/api/knowledge', methods=['GET'])
@login_required
def get_knowledge():
    search = request.args.get('search', '')
    category = request.args.get('category', '')
    conn = get_db()
    c = conn.cursor()
    query = "SELECT * FROM knowledge_base WHERE 1=1"
    params = []
    if search:
        query += " AND (title ILIKE %s OR content ILIKE %s OR tags ILIKE %s)"
        params.extend([f'%{search}%'] * 3)
    if category:
        query += " AND category = %s"
        params.append(category)
    query += " ORDER BY views DESC, created_at DESC"
    c.execute(query, params)
    items = rows_to_list(c.fetchall(), c)
    conn.close()
    return jsonify(items)


@app.route('/api/knowledge', methods=['POST'])
@login_required
def create_knowledge():
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    data = request.get_json()
    k_id = str(uuid.uuid4())
    conn = get_db()
    c = conn.cursor()
    c.execute("""INSERT INTO knowledge_base (id, title, content, category, tags, author)
                 VALUES (%s, %s, %s, %s, %s, %s)""",
              (k_id, data.get('title'), data.get('content'), data.get('category'),
               data.get('tags'), session.get('user_name')))
    conn.commit()
    conn.close()
    return jsonify({'id': k_id, 'success': True}), 201


# ── SERVICE REQUESTS ──────────────────────────────────────────────────────────
@app.route('/api/service-requests', methods=['GET'])
@login_required
def get_service_requests():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM service_requests ORDER BY created_at DESC")
    reqs = rows_to_list(c.fetchall(), c)
    conn.close()
    return jsonify(reqs)


@app.route('/api/service-requests', methods=['POST'])
@login_required
def create_service_request():
    data = request.get_json()
    req_id = str(uuid.uuid4())
    conn = get_db()
    c = conn.cursor()
    c.execute("""INSERT INTO service_requests (id, title, description, category, priority, status, requester_id)
                 VALUES (%s, %s, %s, %s, %s, 'open', %s)""",
              (req_id, data.get('title'), data.get('description'), data.get('category'),
               data.get('priority', 'medium'), session['user_id']))
    conn.commit()
    conn.close()
    return jsonify({'id': req_id, 'success': True}), 201


@app.route('/api/service-requests/<req_id>', methods=['PUT'])
@login_required
def update_service_request(req_id):
    data = request.get_json()
    conn = get_db()
    c = conn.cursor()
    c.execute("""UPDATE service_requests SET status=%s, assigned_to=%s, resolution=%s,
                 updated_at=CURRENT_TIMESTAMP WHERE id=%s""",
              (data.get('status'), data.get('assigned_to'), data.get('resolution'), req_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ── PERFORMANCE REVIEWS ───────────────────────────────────────────────────────
@app.route('/api/performance', methods=['GET'])
@login_required
def get_performance():
    employee_id = request.args.get('employee_id', '')
    conn = get_db()
    c = conn.cursor()
    query = """SELECT pr.*, e.name as employee_name FROM performance_reviews pr
               LEFT JOIN employees e ON pr.employee_id = e.id"""
    params = []
    if employee_id:
        query += " WHERE pr.employee_id = %s"
        params.append(employee_id)
    query += " ORDER BY pr.created_at DESC"
    c.execute(query, params)
    reviews = rows_to_list(c.fetchall(), c)
    conn.close()
    return jsonify(reviews)


@app.route('/api/performance', methods=['POST'])
@login_required
def create_performance():
    if session.get('user_role') not in ('admin', 'lideranca'):
        return jsonify({'error': 'Acesso negado'}), 403
    data = request.get_json()
    pr_id = str(uuid.uuid4())
    scores = json.dumps(data.get('scores', {}))
    conn = get_db()
    c = conn.cursor()
    c.execute("""INSERT INTO performance_reviews (id, employee_id, reviewer_id, period, scores, overall_score, comments, status)
                 VALUES (%s, %s, %s, %s, %s, %s, %s, %s)""",
              (pr_id, data.get('employee_id'), session['user_id'], data.get('period'),
               scores, data.get('overall_score'), data.get('comments'), data.get('status', 'draft')))
    conn.commit()
    conn.close()
    return jsonify({'id': pr_id, 'success': True}), 201


# ── USERS ─────────────────────────────────────────────────────────────────────
@app.route('/api/users', methods=['GET'])
@login_required
def get_users():
    return jsonify({'error': 'Gerenciamento de usuários agora é feito pelo Valore Hub.'}), 403
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, name, email, role, department, created_at, active FROM users")
    users = rows_to_list(c.fetchall(), c)
    conn.close()
    return jsonify(users)


@app.route('/api/users', methods=['POST'])
@login_required
def create_user():
    return jsonify({'error': 'Gerenciamento de usuários agora é feito pelo Valore Hub.'}), 403
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    data = request.get_json()
    user_id = str(uuid.uuid4())
    hashed = bcrypt.hashpw(data.get('password', 'senha123').encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("""INSERT INTO users (id, name, email, password_hash, role, department)
                     VALUES (%s, %s, %s, %s, %s, %s)""",
                  (user_id, data.get('name'), data.get('email'), hashed,
                   data.get('role', 'user'), data.get('department')))
        conn.commit()
    except psycopg2.errors.UniqueViolation:
        conn.rollback()
        return jsonify({'error': 'Email já cadastrado'}), 400
    finally:
        conn.close()
    return jsonify({'id': user_id, 'success': True}), 201


@app.route('/api/users/<user_id>/password', methods=['PUT'])
@login_required
def change_password(user_id):
    return jsonify({'error': 'Gerenciamento de usuários agora é feito pelo Valore Hub.'}), 403
    if session.get('user_id') != user_id and session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    data = request.get_json()
    hashed = bcrypt.hashpw(data.get('password', '').encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE users SET password_hash=%s WHERE id=%s", (hashed, user_id))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


@app.route('/api/users/<user_id>/toggle', methods=['PUT'])
@login_required
def toggle_user_active(user_id):
    return jsonify({'error': 'Gerenciamento de usuários agora é feito pelo Valore Hub.'}), 403
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    # Protege o próprio admin de se desativar
    if session.get('user_id') == user_id:
        return jsonify({'error': 'Você não pode desativar sua própria conta'}), 400
    data = request.get_json()
    active = 1 if data.get('active') else 0
    conn = get_db()
    c = conn.cursor()
    c.execute("UPDATE users SET active=%s WHERE id=%s", (active, user_id))
    conn.commit()
    conn.close()
    acao = 'ACTIVATE' if active else 'DEACTIVATE'
    log_action(session['user_id'], acao, 'user', user_id)
    return jsonify({'success': True})


# ── AUDIT ─────────────────────────────────────────────────────────────────────
@app.route('/api/audit', methods=['GET'])
@login_required
def get_audit():
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    conn = get_db()
    c = conn.cursor()
    c.execute("""
        SELECT al.*, u.name as user_name FROM audit_log al
        LEFT JOIN users u ON al.user_id = u.id
        ORDER BY al.created_at DESC LIMIT 200
    """)
    logs = rows_to_list(c.fetchall(), c)
    conn.close()
    return jsonify(logs)


# ── IMPORT / EXPORT COLABORADORES ────────────────────────────────────────────
_EMPLOYEE_FIELDS = [
    'name', 'email', 'department', 'position', 'manager',
    'admission_date', 'birth_date', 'phone', 'cpf', 'status', 'notes'
]

@app.route('/api/employees/export', methods=['GET'])
@login_required
def export_employees():
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT * FROM employees ORDER BY name ASC")
    employees = rows_to_list(c.fetchall(), c)
    conn.close()

    MAPA = [
        ('name','Nome'),('email','E-mail'),('department','Departamento'),
        ('position','Cargo'),('manager','Gestor'),('admission_date','Admissao'),
        ('birth_date','Nascimento'),('phone','Telefone'),('cpf','CPF'),
        ('status','Status'),('exit_date','Data de Saída'),
        ('is_parent','Pai_Mae'),('parent_type','Tipo_Pai_Mae'),
        ('notes','Observacoes'),
    ]
    output = _io.StringIO()
    writer = _csv.writer(output, delimiter=';')
    writer.writerow([col for _,col in MAPA])
    for emp in employees:
        linha = []
        for field,_ in MAPA:
            val = emp.get(field, '') or ''
            if isinstance(val, bool):
                val = 'Sim' if val else 'Nao'
            linha.append(val)
        writer.writerow(linha)

    csv_bytes = output.getvalue().encode('utf-8-sig')
    return _Response(
        csv_bytes,
        mimetype='text/csv',
        headers={
            'Content-Disposition': 'attachment; filename=colaboradores.csv',
            'Content-Type': 'text/csv; charset=utf-8'
        }
    )

@app.route('/api/employees/import', methods=['POST'])
@login_required
def import_employees():
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403

    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'Nenhum arquivo enviado'}), 400

    filename = file.filename.lower()
    if not (filename.endswith('.csv') or filename.endswith('.xlsx') or filename.endswith('.xls')):
        return jsonify({'error': 'Formato invalido. Use CSV ou XLSX.'}), 400

    def norm(v):
        if v is None:
            return ''
        if hasattr(v, 'strftime'):
            return v.strftime('%Y-%m-%d')
        return str(v).strip()

    def get_field(row, *keys):
        for k in keys:
            val = row.get(k)
            if val is not None and str(val).strip() not in ('', 'None'):
                return norm(val)
        return ''

    try:
        if filename.endswith('.csv'):
            content = file.read().decode('utf-8-sig')
            reader = _csv.DictReader(_io.StringIO(content))
            rows = list(reader)
        else:
            import openpyxl
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                rows.append(dict(zip(headers, row)))

        conn = get_db()
        c = conn.cursor()
        inserted = 0
        skipped = 0
        novos_para_hub = []

        for row in rows:
            name  = get_field(row, 'name', 'Nome', 'nome')
            email = get_field(row, 'email', 'E-mail', 'Email', 'e-mail')

            if not name or not email:
                skipped += 1
                continue

            c.execute("SELECT id FROM employees WHERE email = %s", (email,))
            if c.fetchone():
                skipped += 1
                continue

            emp_id = str(uuid.uuid4())
            c.execute("""
                INSERT INTO employees
                  (id, name, email, department, position, manager,
                   admission_date, birth_date, phone, cpf, status, notes)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """, (
                emp_id,
                name,
                email,
                get_field(row, 'department', 'Departamento', 'departamento'),
                get_field(row, 'position', 'Cargo', 'cargo'),
                get_field(row, 'manager', 'Gestor', 'gestor'),
                get_field(row, 'admission_date', 'Admiss\u00e3o', 'Admissao', 'data_admissao') or None,
                get_field(row, 'birth_date', 'Nascimento', 'data_nascimento') or None,
                get_field(row, 'phone', 'Telefone', 'telefone'),
                get_field(row, 'cpf', 'CPF'),
                get_field(row, 'status', 'Status') or 'active',
                get_field(row, 'notes', 'Observa\u00e7\u00f5es', 'Observacoes', 'observacoes'),
            ))
            novos_para_hub.append((
                name, email,
                get_field(row, 'department', 'Departamento', 'departamento'),
                get_field(row, 'position', 'Cargo', 'cargo'),
            ))
            inserted += 1

        conn.commit()
        conn.close()
        log_action(session['user_id'], 'IMPORT', 'employee', None, f'{inserted} importados')
        for nome_novo, email_novo, dept_novo, cargo_novo in novos_para_hub:
            provisionar_usuario_hub(nome_novo, email_novo, dept_novo, cargo_novo)
        return jsonify({'success': True, 'inserted': inserted, 'skipped': skipped})

    except Exception as e:
        return jsonify({'error': f'Erro ao processar arquivo: {str(e)}'}), 500


@app.route('/api/employees/<emp_id>/permanent', methods=['DELETE'])
@login_required
def delete_employee_permanent(emp_id):
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    conn = get_db()
    c = conn.cursor()
    c.execute("DELETE FROM employees WHERE id = %s", (emp_id,))
    conn.commit()
    conn.close()
    log_action(session['user_id'], 'DELETE_PERMANENT', 'employee', emp_id)
    return jsonify({'success': True})


# ── DISC ──────────────────────────────────────────────────────────────────────
@app.route('/api/employees/<emp_id>/disc', methods=['GET'])
@login_required
def get_disc(emp_id):
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("SELECT * FROM disc_results WHERE employee_id = %s ORDER BY created_at DESC LIMIT 1", (emp_id,))
        row = c.fetchone()
        result = row_to_dict(row, c) if row else None
    except:
        result = None
    conn.close()
    return jsonify(result)

@app.route('/api/employees/<emp_id>/disc', methods=['POST'])
@login_required
def save_disc(emp_id):
    if session.get('user_role') not in ('admin', 'lideranca'):
        return jsonify({'error': 'Acesso negado'}), 403
    data = request.get_json()
    answers = data.get('answers', [])
    # answers é lista de strings: 'D', 'I', 'S' ou 'C'
    scores = {'D': 0, 'I': 0, 'S': 0, 'C': 0}
    for ans in answers:
        if ans in scores:
            scores[ans] += 1
    total = sum(scores.values()) or 1
    pct = {k: round(v/total*100) for k, v in scores.items()}
    dominant = max(scores, key=scores.get)
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id FROM disc_results WHERE employee_id = %s", (emp_id,))
    existing = c.fetchone()
    if existing:
        c.execute(
            "UPDATE disc_results SET d_score=%s, i_score=%s, s_score=%s, c_score=%s, dominant_profile=%s, answers=%s, updated_at=CURRENT_TIMESTAMP WHERE employee_id=%s",
            (pct['D'], pct['I'], pct['S'], pct['C'], dominant, json.dumps(answers), emp_id)
        )
    else:
        c.execute(
            "INSERT INTO disc_results (id, employee_id, d_score, i_score, s_score, c_score, dominant_profile, answers) VALUES (%s,%s,%s,%s,%s,%s,%s,%s)",
            (str(uuid.uuid4()), emp_id, pct['D'], pct['I'], pct['S'], pct['C'], dominant, json.dumps(answers))
        )
    conn.commit()
    conn.close()
    log_action(session['user_id'], 'DISC', 'employee', emp_id)
    return jsonify({'success': True, 'scores': pct, 'dominant': dominant})

# ── CAREER HISTORY ─────────────────────────────────────────────────────────────
@app.route('/api/employees/<emp_id>/career', methods=['GET'])
@login_required
def get_career(emp_id):
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute("SELECT * FROM career_history WHERE employee_id = %s ORDER BY created_at DESC", (emp_id,))
        rows = rows_to_list(c.fetchall(), c)
    except:
        rows = []
    conn.close()
    return jsonify(rows)

@app.route('/api/employees/<emp_id>/career', methods=['POST'])
@login_required
def add_career(emp_id):
    data = request.get_json()
    rec_id = str(uuid.uuid4())
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute(
            "INSERT INTO career_history (id, employee_id, event_type, previous_position, new_position, previous_department, new_department, event_date, notes) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)",
            (rec_id, emp_id, data.get('event_type'), data.get('previous_position'), data.get('new_position'), data.get('previous_department'), data.get('new_department'), data.get('event_date'), data.get('notes'))
        )
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500
    conn.close()
    return jsonify({'success': True, 'id': rec_id})


# ── POLÍTICAS DE EMPRESA ─────────────────────────────────────────────────────
@app.route('/api/policies', methods=['GET'])
@login_required
def get_policies():
    conn = get_db()
    c = conn.cursor()
    if session.get('user_role') == 'admin':
        c.execute("SELECT id, title, category, content, pdf_filename, created_by, created_at, updated_at, status, priority FROM policies ORDER BY category, title")
    else:
        c.execute("SELECT id, title, category, content, pdf_filename, created_by, created_at, updated_at, status, priority FROM policies WHERE status = 'ativa' ORDER BY category, title")
    policies = rows_to_list(c.fetchall(), c)
    conn.close()
    return jsonify(policies)

@app.route('/api/policies', methods=['POST'])
@login_required
def create_policy():
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    title = request.form.get('title', '').strip()
    category = request.form.get('category', '').strip()
    content = request.form.get('content', '').strip()
    status = request.form.get('status', 'ativa')
    priority = request.form.get('priority', 'media')
    if not title or not category:
        return jsonify({'error': 'Título e categoria são obrigatórios'}), 400
    pdf_filename = None
    if 'pdf' in request.files:
        pdf = request.files['pdf']
        if pdf.filename:
            ext = os.path.splitext(pdf.filename)[1].lower()
            if ext != '.pdf':
                return jsonify({'error': 'Apenas arquivos PDF são permitidos'}), 400
            pdf_filename = str(uuid.uuid4()) + '.pdf'
            pdf.save(os.path.join('/var/www/valore-RH/uploads', pdf_filename))
    policy_id = str(uuid.uuid4())
    conn = get_db()
    c = conn.cursor()
    try:
        c.execute(
            "INSERT INTO policies (id, title, category, content, pdf_filename, created_by, created_at, status, priority) VALUES (%s, %s, %s, %s, %s, %s, NOW(), %s, %s)",
            (policy_id, title, category, content, pdf_filename, session.get('user_id'), status, priority)
        )
        conn.commit()
        log_action(session.get('user_id'), 'create_policy', policy_id, f'Política criada: {title}')
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500
    conn.close()
    return jsonify({'success': True, 'id': policy_id})

@app.route('/api/policies/<policy_id>', methods=['GET'])
@login_required
def get_policy(policy_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT id, title, category, content, pdf_filename, created_by, created_at, updated_at, status, priority FROM policies WHERE id = %s", (policy_id,))
    row = c.fetchone()
    policy = row_to_dict(row, c) if row else None
    conn.close()
    if not policy:
        return jsonify({'error': 'Política não encontrada'}), 404
    return jsonify(policy)

@app.route('/api/policies/<policy_id>', methods=['PUT'])
@login_required
def update_policy(policy_id):
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    title = request.form.get('title', '').strip()
    category = request.form.get('category', '').strip()
    content = request.form.get('content', '').strip()
    status = request.form.get('status', 'ativa')
    priority = request.form.get('priority', 'media')
    if not title or not category:
        return jsonify({'error': 'Título e categoria são obrigatórios'}), 400
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT pdf_filename FROM policies WHERE id = %s", (policy_id,))
    row = c.fetchone()
    existing = row_to_dict(row, c) if row else None
    if not existing:
        conn.close()
        return jsonify({'error': 'Política não encontrada'}), 404
    pdf_filename = existing.get('pdf_filename')
    if 'pdf' in request.files:
        pdf = request.files['pdf']
        if pdf.filename:
            ext = os.path.splitext(pdf.filename)[1].lower()
            if ext != '.pdf':
                conn.close()
                return jsonify({'error': 'Apenas arquivos PDF são permitidos'}), 400
            if pdf_filename:
                old_path = os.path.join('/var/www/valore-RH/uploads', pdf_filename)
                if os.path.exists(old_path):
                    os.remove(old_path)
            pdf_filename = str(uuid.uuid4()) + '.pdf'
            pdf.save(os.path.join('/var/www/valore-RH/uploads', pdf_filename))
    try:
        c.execute(
            "UPDATE policies SET title=%s, category=%s, content=%s, pdf_filename=%s, status=%s, priority=%s, updated_at=NOW() WHERE id=%s",
            (title, category, content, pdf_filename, status, priority, policy_id)
        )
        conn.commit()
        log_action(session.get('user_id'), 'update_policy', policy_id, f'Política editada: {title}')
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500
    conn.close()
    return jsonify({'success': True})

@app.route('/api/policies/<policy_id>', methods=['DELETE'])
@login_required
def delete_policy(policy_id):
    if session.get('user_role') != 'admin':
        return jsonify({'error': 'Acesso negado'}), 403
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT pdf_filename FROM policies WHERE id = %s", (policy_id,))
    row = c.fetchone()
    existing = row_to_dict(row, c) if row else None
    if not existing:
        conn.close()
        return jsonify({'error': 'Política não encontrada'}), 404
    if existing.get('pdf_filename'):
        pdf_path = os.path.join('/var/www/valore-RH/uploads', existing['pdf_filename'])
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
    try:
        c.execute("DELETE FROM policies WHERE id = %s", (policy_id,))
        conn.commit()
        log_action(session.get('user_id'), 'delete_policy', policy_id, 'Política excluída')
    except Exception as e:
        conn.rollback()
        conn.close()
        return jsonify({'error': str(e)}), 500
    conn.close()
    return jsonify({'success': True})

@app.route('/api/policies/<policy_id>/pdf', methods=['GET'])
@login_required
def get_policy_pdf(policy_id):
    conn = get_db()
    c = conn.cursor()
    c.execute("SELECT pdf_filename, title FROM policies WHERE id = %s", (policy_id,))
    row = c.fetchone()
    policy = row_to_dict(row, c) if row else None
    conn.close()
    if not policy or not policy.get('pdf_filename'):
        return jsonify({'error': 'PDF não encontrado'}), 404
    return send_from_directory('/var/www/valore-RH/uploads', policy['pdf_filename'], as_attachment=False)

# ── SERVE FRONTEND ────────────────────────────────────────────────────────────
@app.route('/', defaults={'path': ''})
@app.route('/<path:path>')
def serve(path):
    if path and os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    return send_from_directory(app.static_folder, 'index.html')


if __name__ == '__main__':
    init_db()
    print("\n✅ Sistema RH iniciado em http://localhost:5000")
    print("📧 Login: admin@empresa.com | Senha: admin123\n")
    app.run(debug=False, port=5000, host='0.0.0.0')


# Rota explícita para reset de senha (SPA fallback)
@app.route('/reset-senha')
def reset_senha():
    from flask import send_from_directory
    return send_from_directory(app.static_folder, 'index.html')
